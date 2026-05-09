"""Build public/templates/setters-markers-template.xlsx.

Roster-driven dropdowns: Setter/Marker validations point at the editable
roster cells via a defined name, so renaming Andy/Barry/... updates the
dropdown lists. 15 mock rows per term (Term 1..4).

Run: python3 scripts-tmp/build-marking-template.py
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = Path("public/templates/setters-markers-template.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

ROSTER = [
    "Andy","Barry","Cecilia","Douglas","Elaine",
    "Fiona","Gerald","Hannah","Imran","Jocelyn",
    "Kenneth","Lina","Marcus","Nadia","Oliver",
    "Priya","Quentin","Rohan","Siti","Tomas",
    "Uma","Vikram","Wendy","Xavier","Yasmin","Zane",
]
ASSESSMENTS = ["WA1","WA2","WA3","Exam"]
STREAMS = ["G3","G2","G1","G3+G2","G3+G2+G1"]
TERMS = ["T1","T2","T3","T4"]

HEADERS = [
    "SN","Term","Assessment","Stream","Level","Subject","Duration",
    "Setter","Marker","Classes",
    "1","2","3","4","5","6","7","8","9","10",
    "Total","Remarks",
]

wb = Workbook()
ws = wb.active
ws.title = "Deployment"

bold = Font(bold=True)
title_font = Font(bold=True, size=14)
banner_font = Font(bold=True, color="FFFFFF")
banner_fill = PatternFill("solid", start_color="1F4E78")
header_fill = PatternFill("solid", start_color="D9E1F2")
roster_hdr_fill = PatternFill("solid", start_color="FFF2CC")
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row = 1
ws.cell(row=row, column=1, value="Setters & Markers Deployment Template — fill in by Term").font = title_font
ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=len(HEADERS))
row += 2

# Roster section
ws.cell(row=row, column=1,
        value="ROSTER — overwrite with your colleagues' real names. The Setter and Marker dropdowns below update automatically.").font = bold
ws.cell(row=row, column=1).fill = roster_hdr_fill
ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=len(HEADERS))
row += 1

ROSTER_COLS = 4
ROSTER_ROWS = -(-len(ROSTER) // ROSTER_COLS)  # ceil
roster_start_row = row
for r in range(ROSTER_ROWS):
    for c in range(ROSTER_COLS):
        i = c * ROSTER_ROWS + r
        if i < len(ROSTER):
            cell = ws.cell(row=row, column=c + 1, value=ROSTER[i])
            cell.border = border
    row += 1
roster_end_row = row - 1
row += 1

# Header row
header_row = row
for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=header_row, column=i, value=h)
    c.font = bold
    c.fill = header_fill
    c.alignment = center
    c.border = border
ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
row += 1

# Mock data: 15 rows per term
LEVELS = ["Sec 1","Sec 2","Sec 3","Sec 4"]
SUBJECTS = [
    ("English Language","1h"),
    ("Mathematics","1h 15m"),
    ("Combined Science (Phy/Chem)","1h"),
    ("Combined Humanities (SS/Hist)","1h 15m"),
    ("Geography","1h"),
    ("History","1h"),
    ("Literature","1h"),
    ("Mother Tongue","1h"),
    ("Art","1h 30m"),
    ("Design & Technology","1h"),
    ("Food & Consumer Education","1h"),
    ("Physical Education","45m"),
    ("Music","45m"),
    ("Character & Citizenship","45m"),
    ("Computing","1h"),
]
EXAM_DURATION = "1h 45m"

def stream_for(term, idx):
    # Mix of streams across rows
    pool = ["G3","G3","G2","G3+G2","G1","G3","G2","G3+G2+G1","G3","G2","G1","G3","G3+G2","G2","G3"]
    return pool[idx % len(pool)]

def classes_for(level, stream, idx):
    yr = level[-1]
    if "G3" in stream and "G2" in stream:
        return f"{yr}A{(idx%2)+1}, {yr}N{(idx%2)+1}"
    if stream == "G2":
        return f"{yr}N{(idx%2)+1}"
    if stream == "G1":
        return f"{yr}T1"
    return f"{yr}A{(idx%3)+1}, {yr}A{((idx+1)%3)+1}"

def counts_for(classes_str):
    parts = [p.strip() for p in classes_str.split(",") if p.strip()]
    base = [38, 36, 34, 35, 32]
    return [base[i % len(base)] for i in range(len(parts))]

sn = 0
sn_rows = []  # track first data row per term for banner
banner_rows = []

term_count = 1
for term in TERMS:
    # banner row
    banner_row = row
    banner_rows.append(banner_row)
    bcell = ws.cell(row=banner_row, column=1, value=f"─── TERM {term_count} ───")
    bcell.font = banner_font
    bcell.fill = banner_fill
    bcell.alignment = center
    ws.merge_cells(start_row=banner_row, end_row=banner_row, start_column=1, end_column=len(HEADERS))
    row += 1

    for i in range(15):
        sn += 1
        level = LEVELS[i % len(LEVELS)]
        subject, duration = SUBJECTS[i % len(SUBJECTS)]
        stream = stream_for(term, i)
        # Exam mostly in Term 4; WA1/2/3 in T1/T2/T3 with some mixing
        if term == "T4":
            assess = "Exam" if i < 12 else "WA3"
            duration = EXAM_DURATION if assess == "Exam" else duration
        elif term == "T1":
            assess = "WA1"
        elif term == "T2":
            assess = "WA2"
        else:
            assess = "WA3"

        setter = ROSTER[(i + term_count) % len(ROSTER)]
        marker_a = ROSTER[(i + term_count + 3) % len(ROSTER)]
        marker_b = ROSTER[(i + term_count + 7) % len(ROSTER)]
        # Some co-marking entries
        marker = f"{marker_a} / {marker_b}" if i % 4 == 0 else marker_a

        classes = classes_for(level, stream, i)
        counts = counts_for(classes)
        remarks = "Co-marked" if i % 4 == 0 else ("G2 variant of G3 paper" if stream == "G2" and i % 3 == 0 else "")

        ws.cell(row=row, column=1, value=sn)
        ws.cell(row=row, column=2, value=term)
        ws.cell(row=row, column=3, value=assess)
        ws.cell(row=row, column=4, value=stream)
        ws.cell(row=row, column=5, value=level)
        ws.cell(row=row, column=6, value=subject)
        ws.cell(row=row, column=7, value=duration)
        ws.cell(row=row, column=8, value=setter)
        ws.cell(row=row, column=9, value=marker)
        ws.cell(row=row, column=10, value=classes)
        for k in range(10):
            v = counts[k] if k < len(counts) else None
            ws.cell(row=row, column=11 + k, value=v)
        # Total formula across columns K..T
        ws.cell(row=row, column=21, value=f"=SUM(K{row}:T{row})")
        ws.cell(row=row, column=22, value=remarks)
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col).border = border
        row += 1
    term_count += 1

first_data_row = banner_rows[0] + 1
last_data_row = row - 1

# Column widths
widths = [4, 6, 12, 12, 10, 32, 10, 14, 22, 22] + [5]*10 + [8, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Notes
row += 1
ws.cell(row=row, column=1, value="How to fill this in").font = bold
row += 1
notes = [
    "• Roster: rename Andy/Barry/… at the top to your real colleagues. The Setter and Marker dropdowns pick from that list automatically.",
    "• Co-setting / co-marking: pick one name then type ' / ' + the second name (e.g. 'Andy / Barry'). Excel will warn but accept it; points are split.",
    "• Term: T1–T4. The dashboard groups deployments by term.",
    "• Assessment: WA1 / WA2 / WA3 / Exam. WAs are 1pt; Exam is full paper (G3=2pt, G2 standalone=1.5pt, G2 variant of G3=1pt, G1=1pt).",
    "• Stream: G3 / G2 / G1, or combos like G3+G2. G2/G1 papers sharing Subject + Year + Department with a G3 paper auto-link as variants.",
    "• Per-class scripts: enter counts in columns 1..10 in the same order as the Classes cell. Total = SUM.",
    "• Banner rows (─── TERM x ───) are ignored by the importer — they're just visual dividers.",
]
for n in notes:
    ws.cell(row=row, column=1, value=n)
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=len(HEADERS))
    row += 1

# --- Lists sheet (hidden) for Assessments / Streams / Terms ---
ws_lists = wb.create_sheet("Lists")
ws_lists["A1"] = "Assessments"; ws_lists["B1"] = "Streams"; ws_lists["C1"] = "Terms"
for i, v in enumerate(ASSESSMENTS, start=2): ws_lists.cell(row=i, column=1, value=v)
for i, v in enumerate(STREAMS,    start=2): ws_lists.cell(row=i, column=2, value=v)
for i, v in enumerate(TERMS,      start=2): ws_lists.cell(row=i, column=3, value=v)
ws_lists.sheet_state = "hidden"

# --- Defined names ---
# Teachers: union of 4 roster columns on Deployment sheet
roster_refs = []
for c in range(1, ROSTER_COLS + 1):
    col = get_column_letter(c)
    roster_refs.append(f"Deployment!${col}${roster_start_row}:${col}${roster_end_row}")
teachers_ref = ",".join(roster_refs)

wb.defined_names["Teachers"]    = DefinedName("Teachers",    attr_text=teachers_ref)
wb.defined_names["Assessments"] = DefinedName("Assessments", attr_text=f"Lists!$A$2:$A${len(ASSESSMENTS)+1}")
wb.defined_names["Streams"]     = DefinedName("Streams",     attr_text=f"Lists!$B$2:$B${len(STREAMS)+1}")
wb.defined_names["Terms"]       = DefinedName("Terms",       attr_text=f"Lists!$C$2:$C${len(TERMS)+1}")

# --- Data validations ---
def add_dv(formula, cols, allow_blank=True, error_style="stop"):
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank, showErrorMessage=(error_style=="stop"))
    dv.error = "Pick from the dropdown or type a custom value."
    dv.errorStyle = error_style
    ws.add_data_validation(dv)
    for col in cols:
        dv.add(f"{col}{first_data_row}:{col}{last_data_row}")

add_dv("=Terms",       ["B"])
add_dv("=Assessments", ["C"])
add_dv("=Streams",     ["D"])
# Setter/Marker: warning style so users can type "Andy / Barry"
add_dv("=Teachers",    ["H","I"], error_style="warning")

# Integer >= 0 for per-class count columns K..T
dv_int = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showErrorMessage=True)
dv_int.error = "Enter a whole number ≥ 0."
ws.add_data_validation(dv_int)
for c in range(11, 21):
    col = get_column_letter(c)
    dv_int.add(f"{col}{first_data_row}:{col}{last_data_row}")

wb.save(OUT)
print(f"Wrote {OUT} ({OUT.stat().st_size} bytes); data rows {first_data_row}..{last_data_row}")
